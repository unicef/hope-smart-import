import csv
from collections.abc import Callable
from itertools import islice
from typing import TYPE_CHECKING, Any, Iterable

import openpyxl

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    from .types import MultiSheetResult, SheetResult, ValueMapper


class SheetNotError(Exception):
    def __init__(self, index_or_name: int | str) -> None:
        super().__init__(index_or_name)


def identity(x: Any) -> Any:
    return x


def _read_worksheet(
    sheet: "Worksheet",
    start_at_row: int,
    has_header: bool,
    value_mapper: Callable[[Any], Any] = identity,
) -> Iterable[dict[str, Any]]:
    rows = sheet.rows

    field_names = None
    generate_field_names = False
    if has_header:
        field_names = [str(cell.value) for cell in next(rows)]
    else:
        generate_field_names = True

    rows = islice(rows, start_at_row, None)
    for row in rows:
        if generate_field_names:
            field_names = [f"column{d + 1}" for d in range(len(row))]

        yield dict(zip(field_names, (value_mapper(cell.value) for cell in row), strict=True))


def open_xls(
    filepath: str,
    *,
    index_or_name: int | str = 0,
    start_at_row: int = 0,
    has_header: bool = True,
    value_mapper: "ValueMapper" = identity,
) -> "SheetResult":
    wb: "Workbook" = openpyxl.load_workbook(filepath)
    match index_or_name:
        case int():
            sheet_index = index_or_name
        case str():
            try:
                sheet_index = wb.sheetnames.index(index_or_name)
            except ValueError:
                raise SheetNotError(index_or_name)
        case _:
            raise SheetNotError(index_or_name)
    try:
        sh: "Worksheet" = wb.worksheets[sheet_index]
    except IndexError:
        raise SheetNotError(sheet_index)
    yield from _read_worksheet(sh, start_at_row=start_at_row, has_header=has_header, value_mapper=value_mapper)


def open_xls_multi(
    filepath: str,
    indices_or_names: list[int | str] = (0,),
    start_at_row: list[int] | int = 0,
    has_header: list[bool] | bool = True,
    value_mapper: "ValueMapper" = identity,
) -> "MultiSheetResult":
    wb: "Workbook" = openpyxl.load_workbook(filepath)
    indices = []
    for i in indices_or_names:
        try:
            indices.append(i if isinstance(i, int) else wb.sheetnames.index(i))
        except ValueError:
            raise SheetNotError(i)
    for si in indices:
        try:
            sh: "Worksheet" = wb.worksheets[si]
        except IndexError:
            raise SheetNotError(si)
        start_at_row = start_at_row if isinstance(start_at_row, int) else start_at_row[si]
        has_header = has_header if isinstance(has_header, bool) else has_header[si]
        yield (
            si,
            _read_worksheet(sh, start_at_row=start_at_row, has_header=has_header, value_mapper=value_mapper),
        )


def open_csv(
    filepath: str,
    *,
    start_at_row: int = 0,
    has_header: bool = False,
    value_mapper: "ValueMapper" = identity,
) -> "SheetResult":
    with open(filepath) as f:
        if has_header:
            reader = csv.DictReader(f)
            reader = islice(reader, start_at_row, None)
            for row in reader:
                yield {k: value_mapper(v) for k, v in row.items()}
        else:
            reader = csv.reader(f)
            reader = islice(reader, start_at_row, None)
            for row in reader:
                field_names = [f"column{d + 1}" for d in range(len(row))]
                yield dict(zip(field_names, map(value_mapper, row), strict=True))
